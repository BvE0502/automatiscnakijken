"""
Streamlit app: Excel-opdrachten automatisch nakijken (met feedback)
------------------------------------------------------------------

Dit is een kant-en-klare Streamlit-app waarmee je:
- Een **docentenbestand** (model) uploadt met een tabblad `antwoorden` en (optioneel) een tabblad `opgave`.
- Eén of meerdere **studentbestanden** uploadt met uitwerkingen in (bijv.) tabblad `opgave`.
- Automatisch vergelijkt op basis van een checklist uit het tabblad `antwoorden`.
- Per student **percentage goed** en **welke vragen fout** ziet.
- Per student een **feedback-Excel** kunt downloaden, plus een **overzichts-CSV**.

Belangrijk over formules (bv. VERT.ZOEKEN):
- Deze app leest **de laatst opgeslagen waarden** uit een Excel-bestand. Laat studenten dus **het bestand openen, alle formules laten berekenen en opslaan** voordat ze uploaden. 
- Technisch: we gebruiken `openpyxl.load_workbook(..., data_only=True)` om de **gecachete** waarde van een formule te lezen (niet de formule zelf). Als Excel de formule nog niet heeft berekend (of het bestand is niet opnieuw opgeslagen), kan de waarde `None` zijn.

Structuur van het tabblad `antwoorden` (docentenbestand):
- Vereist kolommen: 
  - `vraag_id` (unieke naam per vraag, bv. V1, V2, ...),
  - `sheet` (bv. `opgave`),
  - `cel` (bv. `C12` of `A4`),
  - `verwacht` (de juiste waarde zoals die in Excel hoort te staan, tekst/nummer/datum). 
- Optionele kolommen (voor nauwkeurig vergelijken):
  - `type` ∈ {`auto`, `text`, `number`, `date`, `bool`} (default `auto` → de app raadt op basis van `verwacht`),
  - `tolerance` (numeriek; toegestaan verschil voor getallen, bv. 0.01),
  - `case_sensitive` (TRUE/FALSE, alleen voor tekst, default FALSE),
  - `normalize_space` (TRUE/FALSE; als TRUE, trimt en reduceert meervoudige spaties voor tekstvergelijking).
- Tip: Laat `verwacht` leeg als je de **juiste waarde** wilt aflezen uit het **docentenbestand** zelf (tabblad + cel). In dat geval haalt de app de referentiewaarde uit het model.

Voorbeeld `antwoorden`:
| vraag_id | sheet  | cel  | verwacht  | type   | tolerance | case_sensitive | normalize_space |
|---------:|--------|------|-----------|--------|-----------|----------------|-----------------|
| V1       | opgave | C12  | 825,5     | number | 0,1       | FALSE          | TRUE            |
| V2       | opgave | D5   | Alfa BV   | text   |           | FALSE          | TRUE            |
| V3       | opgave | B9   |           | auto   |           |                |                 |

Installatie (lokaal):
1) Maak een virtuele omgeving en installeer packages:
   ```bash
   pip install -r requirements.txt
   ```
   **requirements.txt** (maak dit bestand naast `app.py`):
   ```
   streamlit
   pandas
   openpyxl
   numpy
   ```
2) Start lokaal:
   ```bash
   streamlit run app.py
   ```

Publiceren op **Streamlit Community Cloud**:
1) Zet `app.py` en `requirements.txt` in een **GitHub-repository**.
2) Ga naar https://streamlit.io/cloud → **New app** → kies je repo/branch → `app.py` als bestandsnaam.
3) Klik **Deploy**. Deel de URL met collega’s. (Zorg dat je toegang instelt zoals gewenst; standaard moeten gebruikers inloggen met e-mail.)

Opzet op een (interne) server:
- Zorg dat Python en de requirements aanwezig zijn en start met `streamlit run app.py --server.port 8501 --server.address 0.0.0.0`.
- Publiceer poort 8501 achter reverse proxy (bv. Nginx) indien nodig.

Copyright/licentie: Vrij te gebruiken binnen je onderwijscontext.
"""

from __future__ import annotations
import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# =============================
# Hulpfuncties: Excel inlezen
# =============================

def _to_bytesio(uploaded_file) -> io.BytesIO:
    """Lees een Streamlit UploadedFile in als BytesIO (zodat we het meerdere keren kunnen gebruiken)."""
    data = uploaded_file.read()
    return io.BytesIO(data)


def _normalize_sheet_name(name: str) -> str:
    """Normaliseer werkbladtitel: lowercased, zonder spaties/tekens, voor robuuste matching."""
    if not isinstance(name, str):
        return ""
    return "".join(ch for ch in name.lower().strip() if ch.isalnum())


def find_answers_sheet(model_bytes: io.BytesIO) -> Optional[str]:
    """Zoek het 'antwoorden' tabblad case-insensitief, met een paar synoniemen.
    Accepteert o.a. Antwoorden/ANTWOORDEN, Antwoord, Answers, Answer, AnswerKey, Model, Key.
    Retourneert de exacte naam zoals die in het bestand staat, of None als niet gevonden.
    """
    # Probeer eerst via openpyxl (snel)
    model_bytes.seek(0)
    names: List[str] = []
    try:
        wb = load_workbook(filename=model_bytes, data_only=True, read_only=True)
        names = list(wb.sheetnames)
    except Exception:
        names = []

    # Fallback via pandas.ExcelFile
    if not names:
        model_bytes.seek(0)
        try:
            xls = pd.ExcelFile(model_bytes, engine="openpyxl")
            names = list(xls.sheet_names)
        except Exception:
            names = []

    if not names:
        return None

    norm_map = {_normalize_sheet_name(n): n for n in names}
    wanted_exact = [
        "antwoorden", "antwoord", "answers", "answer", "answerkey", "model", "key",
    ]

    for w in wanted_exact:
        if w in norm_map:
            return norm_map[w]

    # Substring match als laatste redmiddel
    for k, v in norm_map.items():
        if ("antwoord" in k) or ("answer" in k):
            return v

    return None


def load_answers_df(model_bytes: io.BytesIO) -> pd.DataFrame:
    """Lees het antwoorden-tabblad als DataFrame.
    Robuust voor variaties in tabbladnaam en geeft duidelijke foutmelding met gevonden tabbladen.
    Vereist kolommen: vraag_id, sheet, cel, verwacht (verwacht mag leeg zijn).
    """
    # Zoek het juiste tabblad (case-insensitief + synoniemen)
    sheet = find_answers_sheet(model_bytes)
    if not sheet:
        # Haal lijst met bladen voor foutmelding
        model_bytes.seek(0)
        try:
            wb = load_workbook(filename=model_bytes, data_only=True, read_only=True)
            available = ", ".join(wb.sheetnames)
        except Exception:
            available = "onbekend"
        raise ValueError(
            "Tabblad 'antwoorden' niet gevonden. Hernoem het antwoordenblad naar 'antwoorden' (of 'Antwoorden'). "
            f"Gevonden tabbladen: {available}"
        )

    # Lees het gevonden tabblad in pandas
    model_bytes.seek(0)
    df = pd.read_excel(model_bytes, sheet_name=sheet, engine="openpyxl")

    # Normaliseer kolomnamen
    df.columns = [str(c).strip().lower() for c in df.columns]

    required = {"vraag_id", "sheet", "cel", "verwacht"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"Tabblad '{sheet}' mist verplichte kolommen: {', '.join(sorted(missing))}"
        )

    # Optionele kolommen + defaults
    if "type" not in df.columns:
        df["type"] = "auto"
    if "tolerance" not in df.columns:
        df["tolerance"] = np.nan
    if "case_sensitive" not in df.columns:
        df["case_sensitive"] = False
    if "normalize_space" not in df.columns:
        df["normalize_space"] = True

    # Cast booleans netjes (True/False/1/0/"TRUE"/"FALSE")
    df["case_sensitive"] = df["case_sensitive"].apply(_to_bool)
    df["normalize_space"] = df["normalize_space"].apply(_to_bool)

    return df


def _to_bool(x: Any) -> bool:
    if isinstance(x, str):
        x = x.strip().lower()
        if x in {"true", "waar", "yes", "ja"}:
            return True
        if x in {"false", "onwaar", "no", "nee"}:
            return False
    if isinstance(x, (int, float)):
        return bool(x)
    return bool(x) if x is not None else False


def open_workbook(bytes_io: io.BytesIO, data_only: bool = True):
    """Open een Excel workbook met openpyxl. data_only=True → lees geformatteerde/laatst berekende waarden."""
    bytes_io.seek(0)
    return load_workbook(filename=bytes_io, data_only=data_only)


def get_ws_value(ws: Worksheet, cell_ref: str) -> Any:
    """Haal de waarde uit een werkbladcel (bv. 'C12'). Geeft None terug als cel buiten bereik is."""
    try:
        return ws[cell_ref].value
    except Exception:
        return None


# =============================
# Vergelijkingslogica
# =============================

def canonicalize_text(x: Any, *, normalize_space: bool, case_sensitive: bool) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    s = str(x)
    if normalize_space:
        # Strip en collapse meerdere spaties
        s = " ".join(s.strip().split())
    if not case_sensitive:
        s = s.lower()
    return s


def parse_number(x: Any) -> Optional[float]:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)
    s = str(x).strip()
    # Ondersteun NL notatie met komma ("825,5")
    s = s.replace(" ", "")
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def infer_type(expected: Any, declared: str) -> str:
    declared = (declared or "auto").strip().lower()
    if declared != "auto":
        return declared
    # Heuristiek op basis van "verwacht"
    if isinstance(expected, (int, float, np.integer, np.floating)):
        return "number"
    if isinstance(expected, (pd.Timestamp,)):
        return "date"
    if isinstance(expected, (bool,)):
        return "bool"
    # Check of tekst op getal lijkt
    if parse_number(expected) is not None:
        return "number"
    return "text"


def compare_value(
    got: Any,
    expected: Any,
    *,
    value_type: str,
    tolerance: Optional[float],
    case_sensitive: bool,
    normalize_space: bool,
) -> Tuple[bool, str]:
    """Vergelijk studentwaarde met verwacht. Retourneert (is_goed, uitleg)."""
    vt = value_type.lower()

    if vt == "number":
        g = parse_number(got)
        e = parse_number(expected)
        if g is None or e is None:
            return False, f"Kon getal niet interpreteren (student='{got}', verwacht='{expected}')"
        tol = float(tolerance) if tolerance is not None and not pd.isna(tolerance) else 0.0
        ok = abs(g - e) <= tol
        return ok, f"student={g} / verwacht={e} / tol={tol}"

    if vt == "date":
        # Voor simpelheid: vergelijk als tekst na normalisatie van datumrepresentaties
        try:
            g = pd.to_datetime(got) if not pd.isna(got) else pd.NaT
            e = pd.to_datetime(expected) if not pd.isna(expected) else pd.NaT
            ok = (pd.isna(g) and pd.isna(e)) or (not pd.isna(g) and not pd.isna(e) and g.normalize() == e.normalize())
            return ok, f"student={g.date() if not pd.isna(g) else None} / verwacht={e.date() if not pd.isna(e) else None}"
        except Exception:
            return False, f"Kon datum niet interpreteren (student='{got}', verwacht='{expected}')"

    if vt == "bool":
        mapping = {"true": True, "false": False, "waar": True, "onwaar": False, "ja": True, "nee": False}
        def to_bool(v):
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float)) and not pd.isna(v):
                return bool(v)
            s = str(v).strip().lower()
            return mapping.get(s, None)
        g = to_bool(got)
        e = to_bool(expected)
        ok = (g is not None) and (e is not None) and (g == e)
        return ok, f"student={g} / verwacht={e}"

    # text (default)
    g = canonicalize_text(got, normalize_space=normalize_space, case_sensitive=case_sensitive)
    e = canonicalize_text(expected, normalize_space=normalize_space, case_sensitive=case_sensitive)
    ok = g == e
    return ok, f"student='{g}' / verwacht='{e}' / case_sensitive={case_sensitive}"


# =============================
# Scoren per student
# =============================

@dataclass
class ScoreItem:
    vraag_id: str
    sheet: str
    cel: str
    student_waarde: Any
    verwacht: Any
    type: str
    tolerance: Optional[float]
    case_sensitive: bool
    normalize_space: bool
    goed: bool
    uitleg: str


def score_student(
    student_wb,  # openpyxl workbook (data_only=True)
    answers_df: pd.DataFrame,
    model_wb=None,  # openpyxl workbook met model (voor fallback verwacht-waarden)
) -> List[ScoreItem]:
    results: List[ScoreItem] = []

    # Werkblad cache
    ws_cache: Dict[str, Worksheet] = {}

    for _, row in answers_df.iterrows():
        vraag_id = str(row["vraag_id"]).strip()
        sheet = str(row["sheet"]).strip()
        cel = str(row["cel"]).strip()
        declared_type = str(row.get("type", "auto")).strip().lower()
        tolerance = row.get("tolerance", np.nan)
        case_sensitive = _to_bool(row.get("case_sensitive", False))
        normalize_space = _to_bool(row.get("normalize_space", True))

        # Verwacht-waarde: uit kolom of (indien leeg/NaN) uit model-wb lezen
        expected = row.get("verwacht", None)
        if pd.isna(expected) or expected is None or str(expected).strip() == "":
            expected = None
            if model_wb is not None:
                try:
                    if sheet not in model_wb.sheetnames:
                        expected = None
                    else:
                        ws_model = model_wb[sheet]
                        expected = get_ws_value(ws_model, cel)
                except Exception:
                    expected = None

        # Studentwaarde lezen
        try:
            if sheet not in ws_cache:
                if sheet not in student_wb.sheetnames:
                    ws_cache[sheet] = None  # markeer ontbrekend
                else:
                    ws_cache[sheet] = student_wb[sheet]
            ws = ws_cache.get(sheet)
            student_val = get_ws_value(ws, cel) if ws is not None else None
        except Exception:
            student_val = None

        # Type bepalen en vergelijken
        vtype = infer_type(expected, declared_type)
        goed, uitleg = compare_value(
            student_val,
            expected,
            value_type=vtype,
            tolerance=None if pd.isna(tolerance) else float(tolerance),
            case_sensitive=case_sensitive,
            normalize_space=normalize_space,
        )

        results.append(
            ScoreItem(
                vraag_id=vraag_id,
                sheet=sheet,
                cel=cel,
                student_waarde=student_val,
                verwacht=expected,
                type=vtype,
                tolerance=None if pd.isna(tolerance) else float(tolerance),
                case_sensitive=case_sensitive,
                normalize_space=normalize_space,
                goed=goed,
                uitleg=uitleg,
            )
        )

    return results


def to_details_df(items: List[ScoreItem]) -> pd.DataFrame:
    rows = []
    for it in items:
        rows.append(
            {
                "vraag_id": it.vraag_id,
                "sheet": it.sheet,
                "cel": it.cel,
                "student": it.student_waarde,
                "verwacht": it.verwacht,
                "type": it.type,
                "tolerance": it.tolerance,
                "goed": "JA" if it.goed else "NEE",
                "uitleg": it.uitleg,
            }
        )
    df = pd.DataFrame(rows)
    # Sorteer op vraag_id als het nummeriek/alfanumeriek is
    with pd.option_context("mode.chained_assignment", None):
        try:
            df["_sort"] = df["vraag_id"].str.extract(r"(\d+)").astype(float)
            df = df.sort_values(["_sort", "vraag_id"]).drop(columns=["_sort"]) 
        except Exception:
            df = df.sort_values(["vraag_id"]) 
    return df


def pct_correct(items: List[ScoreItem]) -> float:
    if not items:
        return 0.0
    good = sum(1 for x in items if x.goed)
    return round(100.0 * good / len(items), 1)


# =============================
# Exports: feedback Excel & CSV
# =============================

import xlsxwriter  # gebruikt door pandas ExcelWriter(engine="xlsxwriter") voor formatting

def build_feedback_excel(student_name: str, details_df: pd.DataFrame) -> bytes:
    """Genereer een Excel met een 'feedback'-tabblad en eenvoudige conditional formatting."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        details_df.to_excel(writer, index=False, sheet_name="feedback")
        workbook = writer.book
        worksheet = writer.sheets["feedback"]

        # Zoeker naar kolommen
        headers = list(details_df.columns)
        try:
            goed_col = headers.index("goed")
        except ValueError:
            goed_col = None

        # Conditional formatting: kleur JA/NEE
        if goed_col is not None:
            nrows, ncols = details_df.shape
            start_row, start_col = 1, goed_col  # data begint op rij 2 in Excel
            end_row, end_col = nrows, goed_col
            rng = xlsxwriter.utility.xl_range(start_row, start_col, end_row, end_col)

            fmt_ok = workbook.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
            fmt_bad = workbook.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})

            worksheet.conditional_format(rng, {"type": "text", "criteria": "containing", "value": "JA", "format": fmt_ok})
            worksheet.conditional_format(rng, {"type": "text", "criteria": "containing", "value": "NEE", "format": fmt_bad})

        # Autofit kolommen (ruwweg)
        for i, col in enumerate(details_df.columns):
            width = max(10, int(details_df[col].astype(str).map(len).max()) + 2)
            worksheet.set_column(i, i, min(width, 60))

        # Koptekst met studentnaam
        worksheet.write(0, 0, f"Feedback voor: {student_name}")

    return output.getvalue()


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


# =============================
# Streamlit UI
# =============================

st.set_page_config(page_title="Excel Autonakijken", page_icon="✅", layout="wide")

st.title("✅ Excel-opdrachten automatisch nakijken")

st.markdown(
    """
Upload hieronder eerst je **docentenbestand** met het tabblad `antwoorden` en (optioneel) `opgave`. 
Upload daarna één of meerdere **studentbestanden**. De app vergelijkt waarden per `sheet` + `cel` uit je `antwoorden`-tabel.

> **Tip:** Zorg dat studenten hun bestand **opslaan nadat formules berekend zijn** (anders leest de app soms `None`).
    """
)

with st.sidebar:
    st.header("Stap 1 — Upload bestanden")
    model_file = st.file_uploader(
        "Docentenbestand (Excel, met tabblad 'antwoorden')", type=["xlsx", "xlsm"], key="model"
    )
    student_files = st.file_uploader(
        "Studentbestanden (meerdere toegestaan)", type=["xlsx", "xlsm"], accept_multiple_files=True, key="students"
    )

    st.divider()
    st.header("Vergelijk-instellingen (defaults)")
    default_tol = st.number_input("Standaardtolerantie voor getallen (indien leeg)", value=0.0, step=0.01, format="%.2f")
    default_case = st.checkbox("Hoofdlettergevoelig (tekst)", value=False)
    default_norm = st.checkbox("Spaties normaliseren (tekst)", value=True)

if not model_file:
    st.info("⬅️ Upload eerst het docentenbestand in de sidebar.")
    st.stop()

# Lees model als DataFrame (antwoorden) en als workbook (voor eventueel verwachte waarden uit model)
model_bytes = _to_bytesio(model_file)
answers_df = load_answers_df(model_bytes)
model_wb = open_workbook(_to_bytesio(model_file), data_only=True)

# Vul standaardwaarden in als de optionele kolommen leeg zijn
with pd.option_context("mode.chained_assignment", None):
    answers_df["type"] = answers_df["type"].fillna("auto").replace("", "auto")
    answers_df["tolerance"] = answers_df["tolerance"].fillna(default_tol)
    answers_df["case_sensitive"] = answers_df["case_sensitive"].fillna(default_case)
    answers_df["normalize_space"] = answers_df["normalize_space"].fillna(default_norm)

st.subheader("Checklist uit \u2018antwoorden\u2019")
st.dataframe(
    answers_df,
    use_container_width=True,
    hide_index=True,
)

if not student_files:
    st.info("⬅️ Upload nu één of meer studentbestanden in de sidebar.")
    st.stop()

st.subheader("Resultaten per student")

summary_rows = []
all_zip_parts: List[Tuple[str, bytes]] = []  # (filename, bytes)

for up in student_files:
    student_name = up.name
    st.markdown(f"### 👩‍🎓 {student_name}")

    # Open student workbook
    student_wb = open_workbook(_to_bytesio(up), data_only=True)

    # Score
    items = score_student(student_wb, answers_df, model_wb=model_wb)
    details = to_details_df(items)
    pct = pct_correct(items)

    # Toon samenvatting + details
    col1, col2 = st.columns([1, 2])
    with col1:
        st.metric("% goed", f"{pct}%")
        n_good = int((details["goed"] == "JA").sum())
        n_total = int(len(details))
        st.caption(f"Goed: {n_good} / Totaal: {n_total}")
    with col2:
        st.dataframe(details, use_container_width=True, hide_index=True)

    # Exports per student
    fb_bytes = build_feedback_excel(student_name, details)
    st.download_button(
        label="📥 Download feedback (Excel)",
        data=fb_bytes,
        file_name=f"feedback_{student_name.replace('.xlsx','')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{student_name}_xlsx",
    )

    csv_bytes = df_to_csv_bytes(details)
    st.download_button(
        label="📥 Download details (CSV)",
        data=csv_bytes,
        file_name=f"details_{student_name.replace('.xlsx','')}.csv",
        mime="text/csv",
        key=f"dl_{student_name}_csv",
    )

    # Voor overzichts-CSV
    summary_rows.append({"student": student_name, "%_goed": pct, "goed": n_good, "totaal": len(details)})

    # Verzamel voor ZIP (optioneel later)
    all_zip_parts.append((f"feedback_{student_name.replace('.xlsx','')}.xlsx", fb_bytes))
    all_zip_parts.append((f"details_{student_name.replace('.xlsx','')}.csv", csv_bytes))

# Totaaloverzicht
summary_df = pd.DataFrame(summary_rows).sort_values(["%_goed", "student"], ascending=[False, True])
st.markdown("### 📊 Overzicht alle studenten")
st.dataframe(summary_df, use_container_width=True, hide_index=True)

st.download_button(
    label="📥 Download overzicht (CSV)",
    data=df_to_csv_bytes(summary_df),
    file_name="overzicht_scores.csv",
    mime="text/csv",
    key="dl_overzicht_csv",
)

# (Optioneel) alles in één ZIP
import zipfile

def build_zip(parts: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, b in parts:
            zf.writestr(fname, b)
    return buf.getvalue()

if len(all_zip_parts) > 0:
    st.download_button(
        label="📦 Download alles als ZIP",
        data=build_zip(all_zip_parts),
        file_name="alle_feedback_en_details.zip",
        mime="application/zip",
        key="dl_zip_all",
    )

# =============================
# Veelvoorkomende valkuilen & tips
# =============================
with st.expander("ℹ️ Tips & valkuilen"):
    st.markdown(
        """
- **Formules geven `None`**: Laat studenten het bestand **openen en opslaan**, zodat Excel de formules herberekent en bewaart.
- **Scheidingsteken**: Deze app ondersteunt komma als decimaalteken (NL). Gebruik eventueel `tolerance` om afrondingsverschillen te absorberen.
- **Tekst-vergelijking**: Standaard niet hoofdlettergevoelig en normaliseert spaties; pas aan via kolommen of de sidebar.
- **Verwacht leeg**: Als `verwacht` leeg is, probeert de app de referentiewaarde uit het **modelbestand** te lezen op `sheet` + `cel`.
- **Extra checks**: Voeg rijen toe aan `antwoorden` voor elke cel die je wilt controleren — dus ook tussenresultaten.
- **Datumvelden**: Worden op dag-niveau vergeleken; wil je tijd meenemen, pas de functie `compare_value` aan.
        """
    )
